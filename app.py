# app.py
import io
from typing import Dict, Optional, List, Tuple

import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st


# ----------------------------
# Config
# ----------------------------
MILESTONES = ["CEP", "CGI", "CLL", "VDL", "VAD", "CDD", "CGO", "CER"]
ORDERED_MILESTONES = ["CEP", "CGI", "CLL", "VDL", "VAD", "CDD", "CGO", "CER"]
SEGMENTS = [
    ("CEP", "CGI"),
    ("CGI", "CLL"),
    ("CLL", "VDL"),
    ("VDL", "VAD"),
    ("VAD", "CDD"),
    ("CDD", "CGO"),
    ("CGO", "CER"),
]

P44_FALLBACKS = {"VDL": "VDL_P44", "VAD": "VAD_P44"}

REQUIRED_BASE_COLS = [
    "TENANT_NAME",
    "MASTER_SHIPMENT_ID",
    "POL",
    "POD",
    "CARRIER_NAME",
    "CARRIER_SCAC",
]

DISPLAY_COLS = {
    "TENANT_NAME": "Tenant Name",
    "LANE": "Lane",
    "CARRIER_NAME": "Carrier Name",
    "CARRIER_SCAC": "Carrier SCAC",
    "VOLUME": "Volume (Shipments)",
    "TOTAL_H": "Total Lead Time (Hours)",
    "TOTAL_D": "Total Lead Time (Days)",
    "MIN_H": "Min Lead Time (Hours)",
    "MIN_D": "Min Lead Time (Days)",
    "MED_H": "Median Lead Time (Hours)",
    "MED_D": "Median Lead Time (Days)",
    "PCT_H": "P{p} Lead Time (Hours)",
    "PCT_D": "P{p} Lead Time (Days)",
    "MAX_H": "Max Lead Time (Hours)",
    "MAX_D": "Max Lead Time (Days)",
}

DEFAULT_PERCENTILE_VOLUME_THRESHOLD_PCT = 0.0
DEFAULT_RECOMMENDATION_VOLUME_THRESHOLD_PCT = 0.0


# ----------------------------
# File ingest
# ----------------------------
def _read_input(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()

    if name.endswith(".csv"):
        encodings = ["utf-8", "utf-8-sig", "cp1252", "latin-1"]
        last_err = None
        for enc in encodings:
            try:
                uploaded_file.seek(0)
                return pd.read_csv(uploaded_file, encoding=enc)
            except UnicodeDecodeError as e:
                last_err = e
                continue
        raise ValueError(
            f"Unable to decode CSV with {encodings}. Last error: {last_err}. "
            f"Try exporting as UTF-8, or upload Excel instead."
        )

    if name.endswith(".xlsx") or name.endswith(".xls"):
        uploaded_file.seek(0)
        return pd.read_excel(uploaded_file)

    raise ValueError("Unsupported file type. Please upload a CSV or Excel file.")


# ----------------------------
# Validation
# ----------------------------
def get_missing_columns(
    df: pd.DataFrame,
    start_ms: str,
    end_ms: str,
    whole_journey: bool,
) -> List[str]:
    missing = []

    for col in REQUIRED_BASE_COLS:
        if col not in df.columns:
            missing.append(col)

    selected_milestones = {start_ms, end_ms}
    if whole_journey:
        selected_milestones.update(ORDERED_MILESTONES)

    for ms in sorted(selected_milestones):
        if ms in P44_FALLBACKS:
            fb = P44_FALLBACKS[ms]
            if ms not in df.columns and fb not in df.columns:
                missing.append(f"{ms} (or {fb})")
        else:
            if ms not in df.columns:
                missing.append(ms)

    return missing


# ----------------------------
# Datetime helpers
# ----------------------------
def _coerce_datetimes(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    for c in cols:
        if c in df.columns:
            s = pd.to_datetime(df[c], errors="coerce", utc=True)
            df[c] = s.dt.tz_convert(None)
    return df


def _resolve_timestamp(row: pd.Series, milestone: str) -> pd.Timestamp:
    val = row.get(milestone, pd.NaT)
    if pd.notna(val):
        return val
    fb = P44_FALLBACKS.get(milestone)
    if fb:
        return row.get(fb, pd.NaT)
    return pd.NaT


def _round_hours(x: Optional[float]) -> Optional[float]:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None
    return float(np.round(x, 2))


def _round_days_from_hours(x_hours: Optional[float]) -> Optional[int]:
    if x_hours is None or (isinstance(x_hours, float) and np.isnan(x_hours)):
        return None
    return int(np.round(x_hours / 24.0))


def _safe_quantile(values: pd.Series, q: float) -> Optional[float]:
    values = values.dropna()
    if values.empty:
        return None
    return float(values.quantile(q, interpolation="linear"))


def _make_lane(pol, pod) -> str:
    pol_s = "" if pd.isna(pol) else str(pol)
    pod_s = "" if pd.isna(pod) else str(pod)
    return f"{pol_s} → {pod_s}"


def _segment_col_name(start_ms: str, end_ms: str) -> str:
    return f"SEG_{start_ms}_{end_ms}_HOURS"


def _segment_label(start_ms: str, end_ms: str) -> str:
    return f"{start_ms}-{end_ms}"


def _journey_col_name() -> str:
    return "JOURNEY_LEAD_HOURS"


def _pct_to_count(total_shipments: int, threshold_pct: float) -> int:
    """
    Convert percentage threshold to minimum shipment count.
    Example:
      total = 1000, pct = 10 -> 100
      total = 83, pct = 10 -> 9
    """
    if total_shipments <= 0 or threshold_pct <= 0:
        return 0
    return int(np.ceil(total_shipments * (threshold_pct / 100.0)))


# ----------------------------
# Raw export prep
# ----------------------------
def add_shipment_month_year(raw_df: pd.DataFrame) -> pd.DataFrame:
    df = raw_df.copy()

    dt_cols = ORDERED_MILESTONES + list(P44_FALLBACKS.values())
    dt_cols = list(dict.fromkeys(dt_cols))
    df = _coerce_datetimes(df, dt_cols)

    for ms in ORDERED_MILESTONES:
        df[f"_RES_{ms}"] = df.apply(lambda r: _resolve_timestamp(r, ms), axis=1)
        df[f"_RES_{ms}"] = pd.to_datetime(df[f"_RES_{ms}"], errors="coerce", utc=True).dt.tz_convert(None)

    first_ts = pd.Series(pd.NaT, index=df.index, dtype="datetime64[ns]")
    for ms in ORDERED_MILESTONES:
        first_ts = first_ts.fillna(df[f"_RES_{ms}"])

    df["SHIPMENT_MONTH_YEAR"] = first_ts.dt.strftime("%b %Y").fillna("")
    df = df.drop(columns=[f"_RES_{ms}" for ms in ORDERED_MILESTONES], errors="ignore")
    return df


# ----------------------------
# Core shipment computation
# ----------------------------
def compute_shipment_leadtimes(
    raw: pd.DataFrame,
    start_ms: str,
    end_ms: str,
    shipment_agg: str,
    whole_journey: bool,
) -> pd.DataFrame:
    df = raw.copy()

    dt_cols = [start_ms, end_ms] + ORDERED_MILESTONES + list(P44_FALLBACKS.values())
    dt_cols = list(dict.fromkeys(dt_cols))
    df = _coerce_datetimes(df, dt_cols)

    for ms in ORDERED_MILESTONES:
        df[f"_RES_{ms}"] = df.apply(lambda r: _resolve_timestamp(r, ms), axis=1)
        df[f"_RES_{ms}"] = pd.to_datetime(df[f"_RES_{ms}"], errors="coerce", utc=True).dt.tz_convert(None)

    df["_START_TS"] = df[f"_RES_{start_ms}"]
    df["_END_TS"] = df[f"_RES_{end_ms}"]

    group_cols = ["TENANT_NAME", "POL", "POD", "CARRIER_NAME", "CARRIER_SCAC", "MASTER_SHIPMENT_ID"]

    if whole_journey:
        all_present_mask = pd.Series(True, index=df.index)
        for ms in ORDERED_MILESTONES:
            all_present_mask = all_present_mask & pd.notna(df[f"_RES_{ms}"])

        ordered_mask = pd.Series(True, index=df.index)
        for a, b in SEGMENTS:
            ordered_mask = ordered_mask & (df[f"_RES_{b}"] >= df[f"_RES_{a}"])

        selected_mask = pd.notna(df["_START_TS"]) & pd.notna(df["_END_TS"]) & (df["_END_TS"] >= df["_START_TS"])

        df["_WHOLE_VALID"] = all_present_mask & ordered_mask & selected_mask
        qdf = df[df["_WHOLE_VALID"]].copy()

        if qdf.empty:
            return pd.DataFrame(
                columns=group_cols + ["LANE", _journey_col_name()] + [_segment_col_name(a, b) for a, b in SEGMENTS]
            )

        qdf[_journey_col_name()] = (qdf["_END_TS"] - qdf["_START_TS"]).dt.total_seconds() / 3600.0

        segment_cols = []
        for a, b in SEGMENTS:
            col = _segment_col_name(a, b)
            qdf[col] = (qdf[f"_RES_{b}"] - qdf[f"_RES_{a}"]).dt.total_seconds() / 3600.0
            segment_cols.append(col)

        agg_fn = "min" if shipment_agg.lower().startswith("ear") else "max"
        ship = qdf.groupby(group_cols, dropna=False)[[_journey_col_name()] + segment_cols].agg(agg_fn).reset_index()
        ship["LANE"] = ship.apply(lambda r: _make_lane(r["POL"], r["POD"]), axis=1)
        return ship

    df["_QUALIFIED"] = pd.notna(df["_START_TS"]) & pd.notna(df["_END_TS"]) & (df["_END_TS"] >= df["_START_TS"])
    qdf = df[df["_QUALIFIED"]].copy()

    if qdf.empty:
        return pd.DataFrame(columns=group_cols + ["LANE", _journey_col_name()])

    qdf[_journey_col_name()] = (qdf["_END_TS"] - qdf["_START_TS"]).dt.total_seconds() / 3600.0
    agg_fn = "min" if shipment_agg.lower().startswith("ear") else "max"
    ship = qdf.groupby(group_cols, dropna=False)[[_journey_col_name()]].agg(agg_fn).reset_index()
    ship["LANE"] = ship.apply(lambda r: _make_lane(r["POL"], r["POD"]), axis=1)
    return ship


# ----------------------------
# Counts
# ----------------------------
def compute_lane_and_carrier_counts(shipment_lt: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if shipment_lt.empty:
        lane_counts = pd.DataFrame(columns=["Tenant Name", "Lane", "Shipments"])
        carrier_counts = pd.DataFrame(columns=["Tenant Name", "Carrier Name", "Carrier SCAC", "Shipments"])
        return lane_counts, carrier_counts

    lane_counts = (
        shipment_lt.groupby(["TENANT_NAME", "LANE"], dropna=False)["MASTER_SHIPMENT_ID"]
        .nunique()
        .reset_index()
        .rename(columns={"TENANT_NAME": "Tenant Name", "LANE": "Lane", "MASTER_SHIPMENT_ID": "Shipments"})
        .sort_values(["Shipments", "Lane"], ascending=[False, True])
    )

    carrier_counts = (
        shipment_lt.groupby(["TENANT_NAME", "CARRIER_NAME", "CARRIER_SCAC"], dropna=False)["MASTER_SHIPMENT_ID"]
        .nunique()
        .reset_index()
        .rename(
            columns={
                "TENANT_NAME": "Tenant Name",
                "CARRIER_NAME": "Carrier Name",
                "CARRIER_SCAC": "Carrier SCAC",
                "MASTER_SHIPMENT_ID": "Shipments",
            }
        )
        .sort_values(["Shipments", "Carrier Name"], ascending=[False, True])
    )

    return lane_counts, carrier_counts


def apply_top_n_lanes_filter(shipment_lt: pd.DataFrame, top_n_lanes: int) -> pd.DataFrame:
    if shipment_lt.empty or top_n_lanes <= 0:
        return shipment_lt

    lane_vol = (
        shipment_lt.groupby(["TENANT_NAME", "LANE"], dropna=False)["MASTER_SHIPMENT_ID"]
        .nunique()
        .reset_index()
        .rename(columns={"MASTER_SHIPMENT_ID": "SHIPMENTS"})
    )

    top_lanes = (
        lane_vol.sort_values(["TENANT_NAME", "SHIPMENTS", "LANE"], ascending=[True, False, True])
        .groupby("TENANT_NAME", dropna=False)
        .head(top_n_lanes)[["TENANT_NAME", "LANE"]]
        .drop_duplicates()
    )

    return shipment_lt.merge(top_lanes, on=["TENANT_NAME", "LANE"], how="inner")


# ----------------------------
# Report metrics
# ----------------------------
def _stats_for_series(
    series: pd.Series,
    percentile_p: int,
    include_percentile: bool,
    min_volume_for_percentile: int,
    prefix: str,
) -> Dict[str, Optional[float]]:
    s = series.dropna()
    vol = int(s.shape[0])

    out = {
        f"{prefix}_TOTAL_H": None,
        f"{prefix}_TOTAL_D": None,
        f"{prefix}_MIN_H": None,
        f"{prefix}_MIN_D": None,
        f"{prefix}_MED_H": None,
        f"{prefix}_MED_D": None,
        f"{prefix}_PCT_H": None,
        f"{prefix}_PCT_D": None,
        f"{prefix}_MAX_H": None,
        f"{prefix}_MAX_D": None,
    }

    if vol == 0:
        return out

    total_h = _round_hours(float(s.sum()))
    min_h = _round_hours(float(s.min()))
    med_h = _round_hours(float(s.median()))
    max_h = _round_hours(float(s.max()))

    out[f"{prefix}_TOTAL_H"] = total_h
    out[f"{prefix}_TOTAL_D"] = _round_days_from_hours(total_h)
    out[f"{prefix}_MIN_H"] = min_h
    out[f"{prefix}_MIN_D"] = _round_days_from_hours(min_h)
    out[f"{prefix}_MED_H"] = med_h
    out[f"{prefix}_MED_D"] = _round_days_from_hours(med_h)
    out[f"{prefix}_MAX_H"] = max_h
    out[f"{prefix}_MAX_D"] = _round_days_from_hours(max_h)

    if include_percentile and vol >= int(min_volume_for_percentile):
        pct_h = _round_hours(_safe_quantile(s, percentile_p / 100.0))
        out[f"{prefix}_PCT_H"] = pct_h
        out[f"{prefix}_PCT_D"] = _round_days_from_hours(pct_h)

    return out


def build_duration_configs(start_ms: str, end_ms: str, whole_journey: bool) -> List[Dict[str, str]]:
    configs = [
        {
            "data_col": _journey_col_name(),
            "prefix": "JOURNEY",
            "label": f"{start_ms}-{end_ms}",
            "display_mode": "journey",
        }
    ]

    if whole_journey:
        for a, b in SEGMENTS:
            configs.append(
                {
                    "data_col": _segment_col_name(a, b),
                    "prefix": f"SEG_{a}_{b}",
                    "label": _segment_label(a, b),
                    "display_mode": "segment",
                }
            )

    return configs


def _group_stats(
    g: pd.DataFrame,
    duration_configs: List[Dict[str, str]],
    percentile_p: int,
    include_percentile: bool,
    min_volume_for_percentile: int,
) -> pd.Series:
    result = {"VOLUME": int(g["MASTER_SHIPMENT_ID"].nunique())}

    for cfg in duration_configs:
        result.update(
            _stats_for_series(
                g[cfg["data_col"]],
                percentile_p=percentile_p,
                include_percentile=include_percentile,
                min_volume_for_percentile=min_volume_for_percentile,
                prefix=cfg["prefix"],
            )
        )

    return pd.Series(result)


def build_carrier_lane_report(
    shipment_lt: pd.DataFrame,
    percentile_p: int,
    include_percentile: bool,
    min_volume_for_percentile: int,
    duration_configs: List[Dict[str, str]],
) -> pd.DataFrame:
    base_cols = ["TENANT_NAME", "LANE", "CARRIER_NAME", "CARRIER_SCAC", "VOLUME", "_IS_LANE_ROW", "_POL", "_POD"]

    metric_cols = []
    for cfg in duration_configs:
        pfx = cfg["prefix"]
        metric_cols.extend(
            [
                f"{pfx}_TOTAL_H",
                f"{pfx}_TOTAL_D",
                f"{pfx}_MIN_H",
                f"{pfx}_MIN_D",
                f"{pfx}_MED_H",
                f"{pfx}_MED_D",
                f"{pfx}_PCT_H",
                f"{pfx}_PCT_D",
                f"{pfx}_MAX_H",
                f"{pfx}_MAX_D",
            ]
        )

    cols = base_cols + metric_cols

    if shipment_lt.empty:
        return pd.DataFrame(columns=cols)

    lane_cols = ["TENANT_NAME", "POL", "POD", "LANE"]
    lane_stats = (
        shipment_lt.groupby(lane_cols, dropna=False)
        .apply(lambda g: _group_stats(g, duration_configs, percentile_p, include_percentile, min_volume_for_percentile))
        .reset_index()
    )
    lane_stats["CARRIER_NAME"] = "ALL CARRIERS"
    lane_stats["CARRIER_SCAC"] = ""

    carrier_cols = ["TENANT_NAME", "POL", "POD", "LANE", "CARRIER_NAME", "CARRIER_SCAC"]
    carrier_stats = (
        shipment_lt.groupby(carrier_cols, dropna=False)
        .apply(lambda g: _group_stats(g, duration_configs, percentile_p, include_percentile, min_volume_for_percentile))
        .reset_index()
    )

    lane_stats = lane_stats.sort_values(["TENANT_NAME", "VOLUME", "LANE"], ascending=[True, False, True])

    rows = []
    for _, lr in lane_stats.iterrows():
        tenant, lane, pol, pod = lr["TENANT_NAME"], lr["LANE"], lr["POL"], lr["POD"]

        lane_row = {
            "TENANT_NAME": tenant,
            "LANE": lane,
            "CARRIER_NAME": lr["CARRIER_NAME"],
            "CARRIER_SCAC": lr["CARRIER_SCAC"],
            "VOLUME": lr["VOLUME"],
            "_IS_LANE_ROW": True,
            "_POL": pol,
            "_POD": pod,
        }
        for mc in metric_cols:
            lane_row[mc] = lr.get(mc)
        rows.append(lane_row)

        csub = carrier_stats[
            (carrier_stats["TENANT_NAME"] == tenant)
            & (carrier_stats["POL"].astype(str) == str(pol))
            & (carrier_stats["POD"].astype(str) == str(pod))
        ].sort_values(["VOLUME", "CARRIER_NAME"], ascending=[False, True])

        for _, cr in csub.iterrows():
            carrier_row = {
                "TENANT_NAME": tenant,
                "LANE": "",
                "CARRIER_NAME": cr["CARRIER_NAME"],
                "CARRIER_SCAC": cr["CARRIER_SCAC"],
                "VOLUME": cr["VOLUME"],
                "_IS_LANE_ROW": False,
                "_POL": pol,
                "_POD": pod,
            }
            for mc in metric_cols:
                carrier_row[mc] = cr.get(mc)
            rows.append(carrier_row)

    return pd.DataFrame(rows, columns=cols)


# ----------------------------
# Insights
# ----------------------------
def build_insight_options(duration_configs: List[Dict[str, str]]) -> Dict[str, Dict[str, str]]:
    return {cfg["label"]: cfg for cfg in duration_configs}


def compute_insights_for_metric(
    shipment_lt: pd.DataFrame,
    metric_cfg: Dict[str, str],
    percentile_p: int,
    percentile_threshold_enabled: bool,
    percentile_threshold_pct: float,
    rec_threshold_enabled: bool,
    rec_threshold_pct: float,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Returns:
      lane_summary_df
      carrier_recommendations_df

    Thresholds are percentage-based within each lane.
    """
    metric_col = metric_cfg["data_col"]

    if shipment_lt.empty or metric_col not in shipment_lt.columns:
        lane_summary = pd.DataFrame(
            columns=[
                "TENANT_NAME",
                "LANE",
                "LANE_SHIPMENTS",
                "LANE_MEDIAN_H",
                "LANE_MEDIAN_D",
                "LANE_PXX_H",
                "LANE_PXX_D",
                "CARRIER_COUNT",
                "PERCENTILE_MIN_SHIPMENTS_REQUIRED",
                "RECOMMENDATION_MIN_SHIPMENTS_REQUIRED",
            ]
        )
        carrier_recs = pd.DataFrame(
            columns=[
                "TENANT_NAME",
                "LANE",
                "CARRIER_NAME",
                "CARRIER_SCAC",
                "SHIPMENTS",
                "CARRIER_SHARE_PCT",
                "CARRIER_MEDIAN_H",
                "CARRIER_MEDIAN_D",
                "CARRIER_PXX_H",
                "CARRIER_PXX_D",
                "MEDIAN_ABS_DEV_H",
                "MEDIAN_ABS_DEV_D",
                "PXX_ABS_DEV_H",
                "PXX_ABS_DEV_D",
                "RANK_IN_LANE",
                "RECOMMENDATION_ELIGIBLE",
                "PERCENTILE_ELIGIBLE",
                "PERCENTILE_MIN_SHIPMENTS_REQUIRED",
                "RECOMMENDATION_MIN_SHIPMENTS_REQUIRED",
            ]
        )
        return lane_summary, carrier_recs

    valid = shipment_lt.dropna(subset=[metric_col]).copy()
    if valid.empty:
        return compute_insights_for_metric(
            pd.DataFrame(columns=shipment_lt.columns),
            metric_cfg,
            percentile_p,
            percentile_threshold_enabled,
            percentile_threshold_pct,
            rec_threshold_enabled,
            rec_threshold_pct,
        )

    lane_base = (
        valid.groupby(["TENANT_NAME", "LANE"], dropna=False)
        .agg(
            LANE_SHIPMENTS=("MASTER_SHIPMENT_ID", "nunique"),
            LANE_MEDIAN_H=(metric_col, "median"),
            CARRIER_COUNT=("CARRIER_NAME", "nunique"),
        )
        .reset_index()
    )

    lane_base["PERCENTILE_MIN_SHIPMENTS_REQUIRED"] = lane_base["LANE_SHIPMENTS"].apply(
        lambda x: _pct_to_count(int(x), float(percentile_threshold_pct)) if percentile_threshold_enabled else 0
    )
    lane_base["RECOMMENDATION_MIN_SHIPMENTS_REQUIRED"] = lane_base["LANE_SHIPMENTS"].apply(
        lambda x: _pct_to_count(int(x), float(rec_threshold_pct)) if rec_threshold_enabled else 0
    )

    lane_pxx_rows = []
    for _, row in lane_base.iterrows():
        sub = valid[(valid["TENANT_NAME"] == row["TENANT_NAME"]) & (valid["LANE"] == row["LANE"])][metric_col]
        if percentile_threshold_enabled and int(row["LANE_SHIPMENTS"]) < int(row["PERCENTILE_MIN_SHIPMENTS_REQUIRED"]):
            lane_pxx = None
        else:
            lane_pxx = _safe_quantile(sub, percentile_p / 100.0)
        lane_pxx_rows.append(lane_pxx)

    lane_base["LANE_PXX_H"] = lane_pxx_rows
    lane_base["LANE_MEDIAN_H"] = lane_base["LANE_MEDIAN_H"].apply(_round_hours)
    lane_base["LANE_MEDIAN_D"] = lane_base["LANE_MEDIAN_H"].apply(_round_days_from_hours)
    lane_base["LANE_PXX_H"] = lane_base["LANE_PXX_H"].apply(_round_hours)
    lane_base["LANE_PXX_D"] = lane_base["LANE_PXX_H"].apply(_round_days_from_hours)

    merged = valid.merge(
        lane_base[["TENANT_NAME", "LANE", "LANE_MEDIAN_H", "LANE_SHIPMENTS", "PERCENTILE_MIN_SHIPMENTS_REQUIRED", "RECOMMENDATION_MIN_SHIPMENTS_REQUIRED"]],
        on=["TENANT_NAME", "LANE"],
        how="left",
    )
    merged["ABS_DEV_H"] = (merged[metric_col] - merged["LANE_MEDIAN_H"]).abs()

    carrier_rows = []
    grouped = merged.groupby(["TENANT_NAME", "LANE", "CARRIER_NAME", "CARRIER_SCAC"], dropna=False)

    for (tenant, lane, carrier_name, carrier_scac), g in grouped:
        shipments = int(g["MASTER_SHIPMENT_ID"].nunique())
        lane_shipments = int(g["LANE_SHIPMENTS"].iloc[0]) if not g.empty else 0
        carrier_share_pct = round((shipments / lane_shipments * 100.0), 2) if lane_shipments > 0 else None

        pct_min_shipments = int(g["PERCENTILE_MIN_SHIPMENTS_REQUIRED"].iloc[0]) if not g.empty else 0
        rec_min_shipments = int(g["RECOMMENDATION_MIN_SHIPMENTS_REQUIRED"].iloc[0]) if not g.empty else 0

        carrier_series = g[metric_col].dropna()
        dev_series = g["ABS_DEV_H"].dropna()

        carrier_median_h = _round_hours(float(carrier_series.median())) if not carrier_series.empty else None
        mad_h = _round_hours(float(dev_series.median())) if not dev_series.empty else None

        percentile_eligible = not percentile_threshold_enabled or shipments >= pct_min_shipments
        recommendation_eligible = not rec_threshold_enabled or shipments >= rec_min_shipments

        if percentile_eligible:
            carrier_pxx_h = _round_hours(_safe_quantile(carrier_series, percentile_p / 100.0)) if not carrier_series.empty else None
            dev_pxx_h = _round_hours(_safe_quantile(dev_series, percentile_p / 100.0)) if not dev_series.empty else None
        else:
            carrier_pxx_h = None
            dev_pxx_h = None

        carrier_rows.append(
            {
                "TENANT_NAME": tenant,
                "LANE": lane,
                "CARRIER_NAME": carrier_name,
                "CARRIER_SCAC": carrier_scac,
                "SHIPMENTS": shipments,
                "CARRIER_SHARE_PCT": carrier_share_pct,
                "CARRIER_MEDIAN_H": carrier_median_h,
                "CARRIER_MEDIAN_D": _round_days_from_hours(carrier_median_h),
                "CARRIER_PXX_H": carrier_pxx_h,
                "CARRIER_PXX_D": _round_days_from_hours(carrier_pxx_h),
                "MEDIAN_ABS_DEV_H": mad_h,
                "MEDIAN_ABS_DEV_D": _round_days_from_hours(mad_h),
                "PXX_ABS_DEV_H": dev_pxx_h,
                "PXX_ABS_DEV_D": _round_days_from_hours(dev_pxx_h),
                "RECOMMENDATION_ELIGIBLE": recommendation_eligible,
                "PERCENTILE_ELIGIBLE": percentile_eligible,
                "PERCENTILE_MIN_SHIPMENTS_REQUIRED": pct_min_shipments,
                "RECOMMENDATION_MIN_SHIPMENTS_REQUIRED": rec_min_shipments,
            }
        )

    carrier_recs = pd.DataFrame(carrier_rows)
    if carrier_recs.empty:
        carrier_recs["RANK_IN_LANE"] = pd.Series(dtype="Int64")
        return lane_base, carrier_recs

    carrier_recs["PXX_ABS_DEV_SORT"] = carrier_recs["PXX_ABS_DEV_H"].fillna(np.inf)
    carrier_recs["MEDIAN_ABS_DEV_SORT"] = carrier_recs["MEDIAN_ABS_DEV_H"].fillna(np.inf)

    eligible_recs = carrier_recs[carrier_recs["RECOMMENDATION_ELIGIBLE"]].copy()
    eligible_recs = eligible_recs.sort_values(
        ["TENANT_NAME", "LANE", "MEDIAN_ABS_DEV_SORT", "PXX_ABS_DEV_SORT", "SHIPMENTS", "CARRIER_NAME"],
        ascending=[True, True, True, True, False, True],
    )
    eligible_recs["RANK_IN_LANE"] = eligible_recs.groupby(["TENANT_NAME", "LANE"], dropna=False).cumcount() + 1

    non_eligible = carrier_recs[~carrier_recs["RECOMMENDATION_ELIGIBLE"]].copy()
    non_eligible["RANK_IN_LANE"] = pd.NA

    out = pd.concat([eligible_recs, non_eligible], ignore_index=True)
    out = out.drop(columns=["PXX_ABS_DEV_SORT", "MEDIAN_ABS_DEV_SORT"], errors="ignore")

    lane_summary = lane_base.sort_values(["LANE_SHIPMENTS", "LANE"], ascending=[False, True])
    carrier_recs = out.sort_values(
        ["TENANT_NAME", "LANE", "RECOMMENDATION_ELIGIBLE", "RANK_IN_LANE", "SHIPMENTS", "CARRIER_NAME"],
        ascending=[True, True, False, True, False, True],
    )

    return lane_summary, carrier_recs


def make_lane_selector_labels(lane_summary: pd.DataFrame) -> Dict[str, Tuple[str, str]]:
    mapping = {}
    if lane_summary.empty:
        return mapping

    for _, row in lane_summary.iterrows():
        label = f"{row['TENANT_NAME']} | {row['LANE']} ({int(row['LANE_SHIPMENTS'])} shipments)"
        mapping[label] = (row["TENANT_NAME"], row["LANE"])
    return mapping


def get_selected_lane_outputs(
    shipment_lt: pd.DataFrame,
    lane_summary: pd.DataFrame,
    carrier_recs: pd.DataFrame,
    selected_tenant: str,
    selected_lane: str,
    metric_cfg: Dict[str, str],
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    metric_col = metric_cfg["data_col"]

    lane_row = lane_summary[(lane_summary["TENANT_NAME"] == selected_tenant) & (lane_summary["LANE"] == selected_lane)].copy()
    lane_carriers = carrier_recs[(carrier_recs["TENANT_NAME"] == selected_tenant) & (carrier_recs["LANE"] == selected_lane)].copy()

    top5 = lane_carriers[lane_carriers["RECOMMENDATION_ELIGIBLE"]].copy()
    top5 = top5.sort_values(["RANK_IN_LANE", "SHIPMENTS", "CARRIER_NAME"], ascending=[True, False, True]).head(5)

    carriers_to_plot = top5["CARRIER_NAME"].tolist()
    ship_subset = shipment_lt[
        (shipment_lt["TENANT_NAME"] == selected_tenant)
        & (shipment_lt["LANE"] == selected_lane)
        & (shipment_lt["CARRIER_NAME"].isin(carriers_to_plot))
    ][["TENANT_NAME", "LANE", "MASTER_SHIPMENT_ID", "CARRIER_NAME", "CARRIER_SCAC", metric_col]].copy()

    ship_subset = ship_subset.rename(columns={metric_col: "LEAD_TIME_HOURS"})
    lane_median_h = lane_row["LANE_MEDIAN_H"].iloc[0] if not lane_row.empty else None
    if lane_median_h is not None and not ship_subset.empty:
        ship_subset["ABS_DEV_H"] = (ship_subset["LEAD_TIME_HOURS"] - lane_median_h).abs()
    else:
        ship_subset["ABS_DEV_H"] = np.nan

    return lane_row, top5, ship_subset


def write_insights_excel(
    lane_summary: pd.DataFrame,
    carrier_recs: pd.DataFrame,
    selected_lane_shipments: pd.DataFrame,
) -> bytes:
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        lane_summary.to_excel(writer, sheet_name="Lane Summary", index=False)
        carrier_recs.to_excel(writer, sheet_name="Carrier Recommendations", index=False)
        selected_lane_shipments.to_excel(writer, sheet_name="Selected Lane Shipments", index=False)

        for ws in writer.book.worksheets:
            for cell in ws[1]:
                cell.font = Font(bold=True)
            for idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(idx)].width = 24

    output.seek(0)
    return output.getvalue()


# ----------------------------
# Export helpers
# ----------------------------
def build_export_rename_map(duration_configs: List[Dict[str, str]], percentile_p: int) -> Dict[str, str]:
    export_cols = {
        "TENANT_NAME": DISPLAY_COLS["TENANT_NAME"],
        "LANE": DISPLAY_COLS["LANE"],
        "CARRIER_NAME": DISPLAY_COLS["CARRIER_NAME"],
        "CARRIER_SCAC": DISPLAY_COLS["CARRIER_SCAC"],
        "VOLUME": DISPLAY_COLS["VOLUME"],
    }

    for cfg in duration_configs:
        pfx = cfg["prefix"]
        label = cfg["label"]

        if cfg["display_mode"] == "journey":
            export_cols[f"{pfx}_TOTAL_H"] = DISPLAY_COLS["TOTAL_H"]
            export_cols[f"{pfx}_TOTAL_D"] = DISPLAY_COLS["TOTAL_D"]
            export_cols[f"{pfx}_MIN_H"] = DISPLAY_COLS["MIN_H"]
            export_cols[f"{pfx}_MIN_D"] = DISPLAY_COLS["MIN_D"]
            export_cols[f"{pfx}_MED_H"] = DISPLAY_COLS["MED_H"]
            export_cols[f"{pfx}_MED_D"] = DISPLAY_COLS["MED_D"]
            export_cols[f"{pfx}_PCT_H"] = DISPLAY_COLS["PCT_H"].format(p=percentile_p)
            export_cols[f"{pfx}_PCT_D"] = DISPLAY_COLS["PCT_D"].format(p=percentile_p)
            export_cols[f"{pfx}_MAX_H"] = DISPLAY_COLS["MAX_H"]
            export_cols[f"{pfx}_MAX_D"] = DISPLAY_COLS["MAX_D"]
        else:
            export_cols[f"{pfx}_TOTAL_H"] = f"{label} Total Lead Time (Hours)"
            export_cols[f"{pfx}_TOTAL_D"] = f"{label} Total Lead Time (Days)"
            export_cols[f"{pfx}_MIN_H"] = f"{label} Min Lead Time (Hours)"
            export_cols[f"{pfx}_MIN_D"] = f"{label} Min Lead Time (Days)"
            export_cols[f"{pfx}_MED_H"] = f"{label} Median Lead Time (Hours)"
            export_cols[f"{pfx}_MED_D"] = f"{label} Median Lead Time (Days)"
            export_cols[f"{pfx}_PCT_H"] = f"{label} P{percentile_p} Lead Time (Hours)"
            export_cols[f"{pfx}_PCT_D"] = f"{label} P{percentile_p} Lead Time (Days)"
            export_cols[f"{pfx}_MAX_H"] = f"{label} Max Lead Time (Hours)"
            export_cols[f"{pfx}_MAX_D"] = f"{label} Max Lead Time (Days)"

    return export_cols


def write_excel_counts(lane_counts: pd.DataFrame, carrier_counts: pd.DataFrame) -> bytes:
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        lane_counts.to_excel(writer, sheet_name="Lane Counts", index=False)
        carrier_counts.to_excel(writer, sheet_name="Carrier Counts", index=False)

        bold = Font(bold=True)
        for sheet_name in ["Lane Counts", "Carrier Counts"]:
            ws = writer.book[sheet_name]
            for cell in ws[1]:
                cell.font = bold
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = 26

    output.seek(0)
    return output.getvalue()


def write_excel_final(
    raw_df: pd.DataFrame,
    report_df: pd.DataFrame,
    duration_configs: List[Dict[str, str]],
    percentile_p: int,
) -> bytes:
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    raw_export = add_shipment_month_year(raw_df)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        raw_export.to_excel(writer, sheet_name="Raw Data", index=False)

        export_rename_map = build_export_rename_map(duration_configs, percentile_p)

        if report_df.empty:
            pd.DataFrame(columns=list(export_rename_map.values())).to_excel(writer, sheet_name="Carrier Lane Lead", index=False)
        else:
            df = report_df.copy()
            lane_flags = df["_IS_LANE_ROW"].astype(bool).to_list()
            df = df.drop(columns=["_IS_LANE_ROW", "_POL", "_POD"], errors="ignore")

            ordered_export_keys = [k for k in export_rename_map.keys() if k in df.columns]
            df = df[ordered_export_keys].rename(columns=export_rename_map)
            df.to_excel(writer, sheet_name="Carrier Lane Lead", index=False)

            ws = writer.book["Carrier Lane Lead"]
            bold_font = Font(bold=True)

            for cell in ws[1]:
                cell.font = bold_font

            lane_col_idx = list(df.columns).index(DISPLAY_COLS["LANE"]) + 1
            for i, is_lane in enumerate(lane_flags, start=2):
                if is_lane:
                    ws.cell(row=i, column=lane_col_idx).font = bold_font

            width_map = {
                DISPLAY_COLS["TENANT_NAME"]: 22,
                DISPLAY_COLS["LANE"]: 28,
                DISPLAY_COLS["CARRIER_NAME"]: 28,
                DISPLAY_COLS["CARRIER_SCAC"]: 14,
                DISPLAY_COLS["VOLUME"]: 18,
                "SHIPMENT_MONTH_YEAR": 18,
            }
            for idx, col_name in enumerate(df.columns, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = width_map.get(col_name, 24)

        ws_raw = writer.book["Raw Data"]
        for cell in ws_raw[1]:
            cell.font = Font(bold=True)

    output.seek(0)
    return output.getvalue()


# ----------------------------
# Streamlit UI
# ----------------------------
st.set_page_config(page_title="Ocean Lead Time Analyzer", layout="wide")
st.title("Ocean Lead Time Analyzer (Carrier + Lane Lead Time)")

st.markdown(
    """
Upload a **CSV or Excel** extract. Pick journey **start** and **end** milestones.  
Output includes:
- **Raw Data**
- **Carrier Lane Lead**
- **Insights**
"""
)

uploaded = st.file_uploader("Upload CSV or Excel", type=["csv", "xlsx", "xls"])
if uploaded is None:
    st.stop()

try:
    raw_df = _read_input(uploaded)
except Exception as e:
    st.error(str(e))
    st.stop()

st.success(f"Loaded {raw_df.shape[0]:,} rows × {raw_df.shape[1]:,} columns")

# Sidebar controls
st.sidebar.header("Journey Settings")

start_ms = st.sidebar.selectbox(
    "Journey start milestone",
    MILESTONES,
    index=MILESTONES.index("VDL"),
)

end_ms = st.sidebar.selectbox(
    "Journey end milestone",
    MILESTONES,
    index=MILESTONES.index("VAD"),
)

shipment_agg_label = st.sidebar.radio(
    "Shipment aggregation (multiple containers per master shipment)",
    options=["Earliest (MIN lead time)", "Latest (MAX lead time)"],
    index=0,
)
shipment_agg = "Earliest" if shipment_agg_label.startswith("Earliest") else "Latest"

whole_journey = st.sidebar.checkbox("Calculate for whole journey", value=False)

st.sidebar.divider()
st.sidebar.header("Lane Filter")

top_n_lanes = st.sidebar.number_input(
    "Limit analysis to Top N lanes by volume (0 = all lanes)",
    min_value=0,
    max_value=1000,
    value=0,
    step=5,
)

st.sidebar.divider()
st.sidebar.header("Insights Settings")

include_percentile = st.sidebar.checkbox("Include additional percentile (PXX)", value=True)

percentile_p = st.sidebar.number_input(
    "Percentile value (e.g., 80)",
    min_value=1,
    max_value=99,
    value=80,
    step=1,
    disabled=not include_percentile,
)

limit_by_volume = st.sidebar.checkbox(
    "Only compute percentile if carrier volume share ≥ threshold (%)",
    value=False,
    disabled=not include_percentile,
)

percentile_volume_threshold_pct = st.sidebar.number_input(
    "Percentile volume threshold (%)",
    min_value=0.0,
    max_value=100.0,
    value=float(DEFAULT_PERCENTILE_VOLUME_THRESHOLD_PCT),
    step=1.0,
    disabled=(not include_percentile) or (not limit_by_volume),
)

recommendation_threshold_enabled = st.sidebar.checkbox(
    "Only generate recommendations if carrier volume share ≥ threshold (%)",
    value=False,
)

recommendation_threshold_pct = st.sidebar.number_input(
    "Recommendation volume threshold (%)",
    min_value=0.0,
    max_value=100.0,
    value=float(DEFAULT_RECOMMENDATION_VOLUME_THRESHOLD_PCT),
    step=1.0,
    disabled=not recommendation_threshold_enabled,
)

# Validation
missing_cols = get_missing_columns(
    df=raw_df,
    start_ms=start_ms,
    end_ms=end_ms,
    whole_journey=whole_journey,
)

if missing_cols:
    st.error("The uploaded file is missing required columns:\n\n- " + "\n- ".join(missing_cols))
    st.stop()

# Compute shipment-level data
try:
    shipment_lt_all = compute_shipment_leadtimes(
        raw=raw_df,
        start_ms=start_ms,
        end_ms=end_ms,
        shipment_agg=shipment_agg,
        whole_journey=whole_journey,
    )
except Exception as e:
    st.error(f"Error computing lead times: {e}")
    st.stop()

shipment_lt = apply_top_n_lanes_filter(shipment_lt_all, int(top_n_lanes))

# Metrics
total_shipments_raw = raw_df["MASTER_SHIPMENT_ID"].nunique() if "MASTER_SHIPMENT_ID" in raw_df.columns else None
eligible_shipments = shipment_lt["MASTER_SHIPMENT_ID"].nunique() if not shipment_lt.empty else 0
coverage = (eligible_shipments / total_shipments_raw * 100.0) if total_shipments_raw else 0.0

c1, c2, c3 = st.columns(3)
c1.metric("Total Shipments (unique MASTER_SHIPMENT_ID in file)", f"{total_shipments_raw:,}" if total_shipments_raw is not None else "N/A")
c2.metric("Eligible Shipments (after current rules)", f"{eligible_shipments:,}")
c3.metric("Coverage vs total file shipments", f"{coverage:.1f}%")

if whole_journey:
    st.info("Whole journey mode is ON. Only shipments with all milestone timestamps are included.")

# Counts
lane_counts, carrier_counts = compute_lane_and_carrier_counts(shipment_lt)

st.subheader("Lane & Carrier Counts (shipment volume)")

lc, cc = st.columns(2)

with lc:
    st.markdown("**Lane Counts**")
    st.caption(f"Unique lanes: {lane_counts.shape[0]:,}")
    st.dataframe(lane_counts if lane_counts.shape[0] <= 25 else lane_counts.head(25), use_container_width=True)
    if lane_counts.shape[0] > 25:
        st.caption("Showing Top 25 lanes by shipment volume.")

with cc:
    st.markdown("**Carrier Counts**")
    st.caption(f"Unique carriers: {carrier_counts.shape[0]:,}")
    st.dataframe(carrier_counts if carrier_counts.shape[0] <= 25 else carrier_counts.head(25), use_container_width=True)
    if carrier_counts.shape[0] > 25:
        st.caption("Showing Top 25 carriers by shipment volume.")

counts_excel = write_excel_counts(lane_counts=lane_counts, carrier_counts=carrier_counts)
st.download_button(
    label="Download Lane + Carrier Counts (Excel)",
    data=counts_excel,
    file_name="lane_and_carrier_counts.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

# Shipment preview
st.subheader("Shipment-level lead times (preview)")
preview_cols = [
    "TENANT_NAME",
    "MASTER_SHIPMENT_ID",
    "POL",
    "POD",
    "LANE",
    "CARRIER_NAME",
    "CARRIER_SCAC",
    _journey_col_name(),
]
if whole_journey:
    preview_cols.extend([_segment_col_name(a, b) for a, b in SEGMENTS])

preview_cols = [c for c in preview_cols if c in shipment_lt.columns]
st.dataframe(shipment_lt[preview_cols].head(200), use_container_width=True)

# Final report
duration_configs = build_duration_configs(start_ms=start_ms, end_ms=end_ms, whole_journey=whole_journey)

# For the carrier-lane report, the threshold remains count-based.
# We convert the current percentage threshold against each lane only in insights.
# So here we keep report percentile threshold as 0 if user selected percentage-based logic.
report_df = build_carrier_lane_report(
    shipment_lt=shipment_lt,
    percentile_p=int(percentile_p),
    include_percentile=bool(include_percentile),
    min_volume_for_percentile=0,
    duration_configs=duration_configs,
)

st.subheader("Carrier Lane Lead (preview)")
preview_report = report_df.drop(columns=["_POL", "_POD", "_IS_LANE_ROW"], errors="ignore")
st.dataframe(preview_report, use_container_width=True)

final_excel = write_excel_final(
    raw_df=raw_df,
    report_df=report_df,
    duration_configs=duration_configs,
    percentile_p=int(percentile_p),
)
st.download_button(
    label="Download Final Excel Report",
    data=final_excel,
    file_name=f"carrier_lane_lead_{start_ms}_to_{end_ms}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

# Insights trigger
if "show_insights" not in st.session_state:
    st.session_state["show_insights"] = False

if st.button("Generate Insights"):
    st.session_state["show_insights"] = True

if st.session_state["show_insights"]:
    st.subheader("Insights")

    insight_options = build_insight_options(duration_configs)
    default_metric_label = f"{start_ms}-{end_ms}" if f"{start_ms}-{end_ms}" in insight_options else list(insight_options.keys())[0]

    selected_metric_label = st.selectbox(
        "Choose journey part for insights",
        options=list(insight_options.keys()),
        index=list(insight_options.keys()).index(default_metric_label),
    )

    selected_metric_cfg = insight_options[selected_metric_label]

    lane_summary_df, carrier_recs_df = compute_insights_for_metric(
        shipment_lt=shipment_lt,
        metric_cfg=selected_metric_cfg,
        percentile_p=int(percentile_p),
        percentile_threshold_enabled=bool(include_percentile and limit_by_volume),
        percentile_threshold_pct=float(percentile_volume_threshold_pct),
        rec_threshold_enabled=bool(recommendation_threshold_enabled),
        rec_threshold_pct=float(recommendation_threshold_pct),
    )

    lane_mapping = make_lane_selector_labels(lane_summary_df)

    if not lane_mapping:
        st.warning("No insight data available for the selected settings.")
    else:
        lane_labels = list(lane_mapping.keys())
        selected_lane_label = st.selectbox(
            "Choose lane",
            options=lane_labels,
            index=0,
        )
        selected_tenant, selected_lane = lane_mapping[selected_lane_label]

        lane_row, top5_df, selected_lane_shipments = get_selected_lane_outputs(
            shipment_lt=shipment_lt,
            lane_summary=lane_summary_df,
            carrier_recs=carrier_recs_df,
            selected_tenant=selected_tenant,
            selected_lane=selected_lane,
            metric_cfg=selected_metric_cfg,
        )

        if lane_row.empty:
            st.warning("No lane summary available for the selected lane.")
        else:
            lane_info = lane_row.iloc[0]
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Lane Shipments", f"{int(lane_info['LANE_SHIPMENTS']):,}")
            c2.metric("Carriers in Lane", f"{int(lane_info['CARRIER_COUNT']):,}")
            c3.metric("Lane Median", f"{lane_info['LANE_MEDIAN_D']} d" if pd.notna(lane_info["LANE_MEDIAN_D"]) else "N/A")
            c4.metric(
                f"Lane P{int(percentile_p)}",
                f"{lane_info['LANE_PXX_D']} d" if pd.notna(lane_info["LANE_PXX_D"]) else "N/A",
            )

            st.caption(
                "Recommendations are ranked by lowest deviation from the lane median, "
                "then lowest percentile deviation, then higher shipment volume. "
                "Thresholds are applied as carrier share % of total shipments in the selected lane."
            )

            display_top5 = top5_df[
                [
                    "RANK_IN_LANE",
                    "CARRIER_NAME",
                    "CARRIER_SCAC",
                    "SHIPMENTS",
                    "CARRIER_SHARE_PCT",
                    "CARRIER_MEDIAN_H",
                    "CARRIER_MEDIAN_D",
                    "CARRIER_PXX_H",
                    "CARRIER_PXX_D",
                    "MEDIAN_ABS_DEV_H",
                    "MEDIAN_ABS_DEV_D",
                    "PXX_ABS_DEV_H",
                    "PXX_ABS_DEV_D",
                ]
            ].copy()

            display_top5 = display_top5.rename(
                columns={
                    "RANK_IN_LANE": "Rank",
                    "CARRIER_NAME": "Carrier Name",
                    "CARRIER_SCAC": "Carrier SCAC",
                    "SHIPMENTS": "Shipments",
                    "CARRIER_SHARE_PCT": "Carrier Share (%)",
                    "CARRIER_MEDIAN_H": "Carrier Median (Hours)",
                    "CARRIER_MEDIAN_D": "Carrier Median (Days)",
                    "CARRIER_PXX_H": f"Carrier P{int(percentile_p)} (Hours)",
                    "CARRIER_PXX_D": f"Carrier P{int(percentile_p)} (Days)",
                    "MEDIAN_ABS_DEV_H": "Median Abs Deviation (Hours)",
                    "MEDIAN_ABS_DEV_D": "Median Abs Deviation (Days)",
                    "PXX_ABS_DEV_H": f"P{int(percentile_p)} Abs Deviation (Hours)",
                    "PXX_ABS_DEV_D": f"P{int(percentile_p)} Abs Deviation (Days)",
                }
            )

            st.markdown("**Top 5 Recommended Carriers**")
            st.dataframe(display_top5, use_container_width=True)

            if not top5_df.empty:
                bar_df = top5_df.copy()
                bar_df["Carrier Label"] = bar_df["CARRIER_NAME"].astype(str) + " (" + bar_df["CARRIER_SCAC"].fillna("").astype(str) + ")"

                fig_bar = px.bar(
                    bar_df.sort_values(["RANK_IN_LANE"]),
                    x="Carrier Label",
                    y="MEDIAN_ABS_DEV_H",
                    hover_data={
                        "SHIPMENTS": True,
                        "CARRIER_SHARE_PCT": True,
                        "CARRIER_MEDIAN_H": True,
                        "CARRIER_PXX_H": True,
                        "PXX_ABS_DEV_H": True,
                        "Carrier Label": False,
                        "MEDIAN_ABS_DEV_H": True,
                    },
                    labels={
                        "Carrier Label": "Carrier",
                        "MEDIAN_ABS_DEV_H": "Median Absolute Deviation (Hours)",
                    },
                    title=f"Deviation Ranking: {selected_metric_label} | {selected_lane}",
                )
                st.plotly_chart(fig_bar, use_container_width=True)

                if not selected_lane_shipments.empty:
                    ship_plot = selected_lane_shipments.copy()
                    carrier_order = top5_df["CARRIER_NAME"].tolist()
                    ship_plot["CARRIER_NAME"] = pd.Categorical(ship_plot["CARRIER_NAME"], categories=carrier_order, ordered=True)
                    ship_plot = ship_plot.sort_values("CARRIER_NAME")

                    lane_median_h = lane_info["LANE_MEDIAN_H"]
                    fig_box = px.box(
                        ship_plot,
                        x="CARRIER_NAME",
                        y="LEAD_TIME_HOURS",
                        points="outliers",
                        labels={
                            "CARRIER_NAME": "Carrier",
                            "LEAD_TIME_HOURS": f"{selected_metric_label} Lead Time (Hours)",
                        },
                        title=f"Lead Time Distribution by Carrier: {selected_metric_label} | {selected_lane}",
                    )
                    if pd.notna(lane_median_h):
                        fig_box.add_hline(
                            y=lane_median_h,
                            line_dash="dash",
                            annotation_text=f"Lane Median: {lane_median_h} h",
                        )
                    st.plotly_chart(fig_box, use_container_width=True)

            insights_excel = write_insights_excel(
                lane_summary=lane_summary_df,
                carrier_recs=carrier_recs_df,
                selected_lane_shipments=selected_lane_shipments,
            )

            safe_metric_label = selected_metric_label.replace(" ", "_").replace(">", "").replace("<", "")
            st.download_button(
                label="Download Insights Excel",
                data=insights_excel,
                file_name=f"insights_{safe_metric_label}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
